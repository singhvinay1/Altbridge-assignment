from typing import Any, Dict, List
import os
import time
import pandas as pd
from .templates import get_template_field_order, get_template_headers
from ..settings import get_output_dir


def write_excel(data_rows: List[Dict[str, Any]], template: Dict[str, Any]) -> str:
    template_id = template.get("templateId", "template")
    ts = time.strftime("%Y%m%d_%H%M%S")
    filename = f"extracted_data_{template_id}_{ts}.xlsx"
    out_dir = get_output_dir()
    out_path = os.path.join(out_dir, filename)
    
    # Check if template supports multi-sheet structure
    if template.get("multiSheet", False) and "sheets" in template:
        write_multi_sheet_excel(data_rows, template, out_path)
    else:
        # Legacy single-sheet format with header
        fields = get_template_field_order(template)
        headers = get_template_headers(template)
        rows = [[row.get(k, "") for k in fields] for row in data_rows]
        df = pd.DataFrame(rows, columns=headers)
        
        # Add header row for both template1 and template2
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        
        # Create a new workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Sheet1"
        
        # Add the header title in row 1
        ws.cell(row=1, column=1, value="Data Extraction Template - Private Equity Funds")
        
        # Style the header cell
        header_font = Font(bold=True, color="FFFFFF", size=14)
        header_fill = PatternFill(start_color="000000", end_color="000000", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        header_cell = ws.cell(row=1, column=1)
        header_cell.font = header_font
        header_cell.fill = header_fill
        header_cell.alignment = header_alignment
        
        # Merge cells for the header across all columns
        ws.merge_cells(f'A1:{chr(65 + len(headers) - 1)}1')
        
        # Add column headers in row 2
        for col_idx, header in enumerate(headers, 1):
            ws.cell(row=2, column=col_idx, value=header)
        
        # Add data starting from row 3
        for row_idx, row_data in enumerate(rows, 3):
            for col_idx, value in enumerate(row_data, 1):
                ws.cell(row=row_idx, column=col_idx, value=value)
        
        # Save the workbook
        wb.save(out_path)
    
    return filename


def write_multi_sheet_excel(data_rows: List[Dict[str, Any]], template: Dict[str, Any], out_path: str) -> None:
    """Write Excel file with multiple sheets based on template structure."""
    with pd.ExcelWriter(out_path, engine='openpyxl') as writer:
        sheets = template.get("sheets", [])
        
        for sheet_config in sheets:
            sheet_name = sheet_config.get("name", "Sheet")
            fields = [field["key"] for field in sheet_config.get("fields", [])]
            headers = [field["header"] for field in sheet_config.get("fields", [])]
            
            # Create sheet data - include all rows but only with fields for this sheet
            sheet_data = []
            for row in data_rows:
                sheet_row = {}
                for field in fields:
                    sheet_row[field] = row.get(field, "")
                sheet_data.append(sheet_row)
            
            # If no data rows, create empty sheet with headers
            if not sheet_data:
                sheet_data = [{}]
            
            # Create DataFrame for this sheet
            rows = [[row.get(k, "") for k in fields] for row in sheet_data]
            df = pd.DataFrame(rows, columns=headers)
            
            # Write to Excel sheet
            df.to_excel(writer, sheet_name=sheet_name, index=False)
            
            # Add description as a comment or note (if possible)
            description = sheet_config.get("description", "")
            if description:
                # For now, we'll add the description as the first row
                # In a more advanced implementation, we could add it as a comment
                pass


