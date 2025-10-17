import json
import os
import re
from typing import Any, Dict, List, Optional
from openpyxl import load_workbook
from settings import get_templates_dir


def slugify_to_key(text: str) -> str:
    s = (text or "").strip().lower()
    s = re.sub(r"[^a-z0-9]+", "_", s)
    s = re.sub(r"_+", "_", s).strip("_")
    return s or "field"


def get_all_template_dirs() -> List[str]:
    dirs: List[str] = []
    # Project templates dir
    dirs.append(get_templates_dir())
    # Optional external dir from env
    external = os.getenv("EXTERNAL_TEMPLATE_DIR")
    if external and os.path.isdir(external):
        dirs.append(external)
    # Specific provided folder if present
    provided = os.path.join(os.path.expanduser("/Users/vinaysingh/Desktop/Ass"), "Intern_task_input files")
    if os.path.isdir(provided):
        dirs.append(provided)
    # De-dup while preserving order
    seen = set()
    uniq: List[str] = []
    for d in dirs:
        if d not in seen:
            seen.add(d)
            uniq.append(d)
    return uniq


def _find_file_in_dirs(filename: str) -> Optional[str]:
    for d in get_all_template_dirs():
        path = os.path.join(d, filename)
        if os.path.exists(path):
            return path
    return None


def _find_xlsx_for_template(template_id: str) -> Optional[str]:
    # Prefer exact known names
    candidates: List[str] = []
    if template_id == "template1":
        candidates = [
            "Extraction Template 1.xlsx",
            "template1.xlsx",
            "Template 1.xlsx",
        ]
    elif template_id == "template2":
        candidates = [
            "Extraction Template 2.xlsx",
            "template2.xlsx",
            "Template 2.xlsx",
        ]

    for name in candidates:
        found = _find_file_in_dirs(name)
        if found:
            return found

    # Fallback: scan for any .xlsx with a matching digit in name
    for d in get_all_template_dirs():
        try:
            for fn in os.listdir(d):
                if not fn.lower().endswith(".xlsx"):
                    continue
                if template_id == "template1" and "1" in fn:
                    return os.path.join(d, fn)
                if template_id == "template2" and "2" in fn:
                    return os.path.join(d, fn)
        except Exception:
            continue
    return None


def _load_template_json(path: str) -> Dict[str, Any]:
    with open(path, "r", encoding="utf-8") as f:
        return json.load(f)


def _load_template_xlsx(path: str, template_id: str) -> Dict[str, Any]:
    # Read first sheet headers as the template columns
    wb = load_workbook(path, read_only=True)
    ws = wb.active
    headers: List[str] = [str(cell.value) for cell in ws[1] if cell.value is not None]
    wb.close()
    fields = [{"key": slugify_to_key(h), "header": h} for h in headers]
    return {
        "templateId": template_id,
        "description": f"Loaded from {os.path.basename(path)}",
        "fields": fields,
    }


def load_template(template_id: str) -> Dict[str, Any]:
    # Try JSON in known dirs
    json_path = _find_file_in_dirs(f"{template_id}.json")
    if json_path:
        return _load_template_json(json_path)

    # Try XSLX filenames
    xlsx_path = _find_xlsx_for_template(template_id)
    if xlsx_path:
        return _load_template_xlsx(xlsx_path, template_id)

    raise FileNotFoundError(f"Template not found for id: {template_id}")


def get_template_field_order(template: Dict[str, Any]) -> list[str]:
    """Get field order for single-sheet templates (legacy support)."""
    if template.get("multiSheet", False) and "sheets" in template:
        # For multi-sheet templates, return all fields from all sheets
        all_fields = []
        for sheet in template.get("sheets", []):
            all_fields.extend([f["key"] for f in sheet.get("fields", [])])
        return all_fields
    return [f["key"] for f in template.get("fields", [])]


def get_template_headers(template: Dict[str, Any]) -> list[str]:
    """Get headers for single-sheet templates (legacy support)."""
    if template.get("multiSheet", False) and "sheets" in template:
        # For multi-sheet templates, return all headers from all sheets
        all_headers = []
        for sheet in template.get("sheets", []):
            all_headers.extend([f["header"] for f in sheet.get("fields", [])])
        return all_headers
    return [f["header"] for f in template.get("fields", [])]


def get_all_template_fields(template: Dict[str, Any]) -> Dict[str, str]:
    """Get all fields from a template as key->header mapping."""
    field_map = {}
    
    if template.get("multiSheet", False) and "sheets" in template:
        for sheet in template.get("sheets", []):
            for field in sheet.get("fields", []):
                field_map[field["key"]] = field["header"]
    else:
        for field in template.get("fields", []):
            field_map[field["key"]] = field["header"]
    
    return field_map


