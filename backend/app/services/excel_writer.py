from typing import Any, Dict, List, Tuple
import os
import time
import io
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from .templates import get_template_field_order, get_template_headers
from ..settings import get_output_dir


def write_excel(data_rows: List[Dict[str, Any]], template: Dict[str, Any]) -> Tuple[str, bytes]:
    template_id = template.get("templateId", "template")
    ts = time.strftime("%Y%m%d_%H%M%S")
    filename = f"extracted_data_{template_id}_{ts}.xlsx"
    out_dir = get_output_dir()
    out_path = os.path.join(out_dir, filename)
    
    # Create workbook in memory
    wb = Workbook()
    
    # Check if template supports multi-sheet structure
    if template.get("multiSheet", False) and "sheets" in template:
        write_multi_sheet_excel_to_workbook(data_rows, template, wb)
    else:
        # Legacy single-sheet format with header
        fields = get_template_field_order(template)
        headers = get_template_headers(template)
        rows = [[row.get(k, "") for k in fields] for row in data_rows]
        
        ws = wb.active
        ws.title = "Sheet1"
        
        # Add the header title in row 1
        ws.cell(row=1, column=1, value="Data Extraction Template - Private Equity Funds")
        
        # Style the header cell
        header_font = Font(bold=True, color="FFFFFF", size=14)
        header_fill = PatternFill(start_color="000000", end_color="000000", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        header_cell = ws.cell(row=1, column=1)
        header_cell.font = header_font
        header_cell.fill = header_fill
        header_cell.alignment = header_alignment
        
        # Merge cells for the header across all columns
        ws.merge_cells(f'A1:{chr(65 + len(headers) - 1)}1')
        
        # Add column headers in row 2
        for col_idx, header in enumerate(headers, 1):
            ws.cell(row=2, column=col_idx, value=header)
        
        # Add data starting from row 3
        for row_idx, row_data in enumerate(rows, 3):
            for col_idx, value in enumerate(row_data, 1):
                ws.cell(row=row_idx, column=col_idx, value=value)
    
    # Save to memory buffer
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    file_content = buffer.getvalue()
    
    # Also save to disk for local development
    try:
        wb.save(out_path)
    except Exception:
        pass  # Ignore file system errors in serverless environments
    
    return filename, file_content


def write_multi_sheet_excel_to_workbook(data_rows: List[Dict[str, Any]], template: Dict[str, Any], wb: Workbook) -> None:
    """Write Excel file with multiple sheets based on template structure."""
    # Remove default sheet
    wb.remove(wb.active)
    
    sheets = template.get("sheets", [])
    
    for sheet_config in sheets:
        sheet_name = sheet_config.get("name", "Sheet")
        fields = [field["key"] for field in sheet_config.get("fields", [])]
        headers = [field["header"] for field in sheet_config.get("fields", [])]
        
        # Create new worksheet
        ws = wb.create_sheet(title=sheet_name)
        
        # Add headers
        for col_idx, header in enumerate(headers, 1):
            ws.cell(row=1, column=col_idx, value=header)
        
        # Add data rows
        for row_idx, row_data in enumerate(data_rows, 2):
            for col_idx, field in enumerate(fields, 1):
                value = row_data.get(field, "")
                ws.cell(row=row_idx, column=col_idx, value=value)
        
        # Add description as a comment or note (if possible)
        description = sheet_config.get("description", "")
        if description:
            # For now, we'll add the description as the first row
            # In a more advanced implementation, we could add it as a comment
            pass


